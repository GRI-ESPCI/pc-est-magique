import reportlab as rl
from reportlab.lib import units, pagesizes, styles
from reportlab.pdfgen import canvas
from reportlab.platypus import flowables
import io
import os
import openpyxl
import zipfile

from app import db
from app.models import ClubQSpectacle, ClubQVoeu, PCeen, ClubQSeason, ClubQSalle
from app.utils.validators import DataRequired
import wtforms
import flask
from app.routes.club_q import forms
from flask_babel import _
from datetime import datetime
from flask import request


cm = rl.lib.units.cm


def pdf_client(pceen, season, voeux):
    # Create a PDF buffer
    buffer = io.BytesIO()

    # Code de génération de facture du Club Q originalement par Loïc Simon - Adapté aux besoin de pem
    canvas = rl.pdfgen.canvas.Canvas(buffer, pagesize=rl.lib.pagesizes.A4)
    canvas.setAuthor("Club Q ESPCI Paris - PSL")
    canvas.setTitle(f"Compte-rendu {pceen.nom}")
    canvas.setSubject(f"Saison {season.nom}")

    styles = rl.lib.styles.getSampleStyleSheet()
    styleN = styles["Normal"]
    canvas.setFont("Times-Bold", 18)

    descr = [
        f"Nom : <b>{pceen.nom}</b>",
        f"Prénom : <b>{pceen.prenom}</b>",
        f"Promo : <b>{pceen.promo}</b>",
        f"Adresse e-mail : <b>{pceen.email}</b>",
    ]

    specs = [
        f"{voeu.spectacle.nom} - {voeu.places_attribuees} place(s) -\t\t {voeu.spectacle.unit_price} € /place"
        for voeu in voeux
    ]

    specs.append("-" * 158)
    specs.append(f"Total à payer : {pceen_prix_total(pceen, voeux)} €.")

    # Conversion en objets reportlab
    blocs_descr = []
    for txt in descr:
        blocs_descr.append(rl.platypus.Paragraph(txt, styleN))
        blocs_descr.append(rl.platypus.Spacer(1, 0.2 * cm))

    blocs_specs = []
    for txt in specs:
        blocs_specs.append(rl.platypus.Paragraph(txt, styleN))
        blocs_specs.append(rl.platypus.Spacer(1, 0.2 * cm))

    # IMPRESSION
    canvas.setFont("Times-Bold", 18)
    canvas.drawString(1 * cm, 28 * cm, f"Compte-rendu Club Q")

    canvas.setFont("Times-Bold", 12)
    canvas.drawString(1 * cm, 27.2 * cm, f"Saison {season.nom}")

    dX, dY = canvas.drawImage(
        os.path.join("app", "static", "img", "espci.png"),
        14 * cm,
        27.22 * cm,
        width=5.63 * cm,
        height=1.2 * cm,
        mask="auto",
    )
    dX, dY = canvas.drawImage(
        os.path.join("app", "static", "img", "club_q.png"),
        11.5 * cm,
        26.8 * cm,
        width=2 * cm,
        height=2 * cm,
        mask="auto",
    )

    frame_descr = rl.platypus.Frame(1 * cm, 23.5 * cm, 19 * cm, 3 * cm, showBoundary=True)
    frame_descr.addFromList(blocs_descr, canvas)

    frame_specs = rl.platypus.Frame(1 * cm, 1 * cm, 19 * cm, 22 * cm, showBoundary=True)
    frame_specs.addFromList(blocs_specs, canvas)

    canvas.showPage()

    # SAUVEGARDE
    canvas.save()

    # Reset the buffer's file pointer to the beginning
    buffer.seek(0)

    return buffer


def pdf_spectacle(spec, season, voeux_attrib, voeux_nan_attrib):
    # Create a PDF buffer
    buffer = io.BytesIO()

    # Code de génération de résumé d'un spectacle du Club Q originalement par Loïc Simon - Adapté aux besoin de pem
    canvas = rl.pdfgen.canvas.Canvas(buffer, pagesize=rl.lib.pagesizes.A4)
    styles = rl.lib.styles.getSampleStyleSheet()
    styleN = styles["Normal"]
    canvas.setFont("Times-Bold", 18)

    descr = [
        f"Titre : <b>{spec.nom}</b> ({spec.categorie or 'catégorie inconnue'})",
        f"Salle : <b>{spec.salle}</b>, {spec.salle.adresse or 'adresse inconnue'}",
        f"Date : <b>{spec.date or '(heure inconnue)'}</b>",
        f"Nombre de places : <b>{spec.nb_tickets}</b> – Demandées : <b>{spec.sum_places_demandees}</b> – Attribuées : <b>{spec.sum_places_attribuees}</b>",
    ]

    eleves = [
        f"{attrib.pceen.full_name}, {attrib.pceen.promo}, {attrib.pceen.email} - {attrib.places_attribuees} place(s)"
        for attrib in voeux_attrib
    ]

    # Conversion en objets reportlab

    blocs_descr = []
    for txt in descr:
        blocs_descr.append(rl.platypus.Paragraph(txt, styleN))
        blocs_descr.append(rl.platypus.Spacer(1, 0.2 * cm))

    blocs_eleves = []
    for txt in eleves:
        blocs_eleves.append(rl.platypus.Paragraph(txt, styleN))
        blocs_eleves.append(rl.platypus.Spacer(1, 0.2 * cm))

    # IMPRESSION PAGE 1
    canvas.setFont("Times-Bold", 18)
    canvas.drawString(1 * cm, 28 * cm, f"Compte-rendu {spec.nom}")

    canvas.setFont("Times-Bold", 12)
    canvas.drawString(1 * cm, 27.2 * cm, f"Saison {season.nom}")

    dX, dY = canvas.drawImage(
        os.path.join("app", "static", "img", "espci.png"),
        14 * cm,
        27.22 * cm,
        width=5.63 * cm,
        height=1.2 * cm,
        mask="auto",
    )
    dX, dY = canvas.drawImage(
        os.path.join("app", "static", "img", "club_q.png"),
        11.5 * cm,
        26.8 * cm,
        width=2 * cm,
        height=2 * cm,
        mask="auto",
    )

    frame_descr = rl.platypus.Frame(1 * cm, 23.5 * cm, 19 * cm, 3 * cm, showBoundary=1)
    frame_descr.addFromList(blocs_descr, canvas)

    frame_eleves = rl.platypus.Frame(1 * cm, 1 * cm, 19 * cm, 22 * cm, showBoundary=1)
    frame_eleves.addFromList(blocs_eleves, canvas)

    canvas.showPage()

    # PAGE 2 (liste d'attente)

    voeux = [voeu for voeu in voeux_nan_attrib]

    if voeux:
        voeux.sort(key=lambda a: a.pceen.discontent or 0, reverse=True)

        voeux_txt = ["Liste d'attente :"]
        for voeu in voeux:
            voeux_txt.append(
                f"M: {round(voeu.pceen.discontent or 0, 2)} - {voeu.pceen.full_name} - {voeu.places_demandees} place(s) ({voeu.places_minimum} min) - Voeu n°{voeu.priorite}"
            )

    else:
        voeux_txt = ["Pas de liste d'attente."]

    blocs_descr = []  # Détruits à la précédente impression...
    for txt in descr:
        blocs_descr.append(rl.platypus.Paragraph(txt, styleN))
        blocs_descr.append(rl.platypus.Spacer(1, 0.2 * cm))

    blocs_voeux = []
    for txt in voeux_txt:
        blocs_voeux.append(rl.platypus.Paragraph(txt, styleN))
        blocs_voeux.append(rl.platypus.Spacer(1, 0.2 * cm))

    # IMPRESSION PAGE 2
    canvas.setFont("Times-Bold", 18)
    canvas.drawString(1 * cm, 28 * cm, f"Liste d'attente {spec.nom}")

    canvas.setFont("Times-Bold", 12)
    canvas.drawString(1 * cm, 27.2 * cm, f"Saison {season.nom}")

    dX, dY = canvas.drawImage(
        os.path.join("app", "static", "img", "espci.png"),
        14 * cm,
        27.22 * cm,
        width=5.63 * cm,
        height=1.2 * cm,
        mask="auto",
    )
    dX, dY = canvas.drawImage(
        os.path.join("app", "static", "img", "club_q.png"),
        11.5 * cm,
        26.8 * cm,
        width=2 * cm,
        height=2 * cm,
        mask="auto",
    )

    frame_descr = rl.platypus.Frame(1 * cm, 23.5 * cm, 19 * cm, 3 * cm, showBoundary=1)
    frame_descr.addFromList(blocs_descr, canvas)

    frame_eleves = rl.platypus.Frame(1 * cm, 1 * cm, 19 * cm, 22 * cm, showBoundary=1)
    frame_eleves.addFromList(blocs_voeux, canvas)

    canvas.showPage()

    # SAUVEGARDE

    canvas.save()

    # Reset the buffer's file pointer to the beginning
    buffer.seek(0)

    return buffer


def excel_spectacle(spec, season, voeux_attrib, voeux_nan_attrib):
    # Code de génération de résumé de spectacle du Club Q originalement par Loïc Simon - Adapté aux besoin de pem
    # Code assez (très) moche, attention - il y a moyen de faire ça beaucoup mieux, mais bon...
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Élèves"

    sheet["A1"] = f"Club Q – Récapitulatif spectacle : {spec.nom}"
    sheet["A1"].font = openpyxl.styles.Font(bold=True)

    sheet["A2"] = "Saison :"
    sheet["B2"] = season.nom
    sheet["D2"] = "Prix du billet :"
    sheet["E2"] = spec.unit_price
    sheet["E2"].number_format = "#,##0.00 €_-"

    sheet["A3"] = "Date et heure :"
    sheet["B3"] = spec.date
    sheet["B3"].number_format = "dd/mm/yy hh:mm"
    sheet["D3"] = "Salle :"
    sheet["E3"] = spec.salle.nom

    sheet["A4"] = "Places proposées :"
    sheet["B4"] = spec.nb_tickets
    sheet["D4"] = "Places vendues :"
    sheet["E4"] = sum((voeu.places_attribuees or 0) for voeu in voeux_attrib)

    sheet["A6"] = "NOM"
    sheet["B6"] = "Prénom"
    sheet["C6"] = "Promo/autre"
    sheet["D6"] = "Adresse mail"
    sheet["E6"] = "Places"
    sheet["F6"] = "Montant"

    npa = 0
    tot = 0
    for i, voeu in enumerate(sorted(voeux_attrib, key=lambda v: v.pceen.full_name)):
        pceen = voeu.pceen
        sheet.cell(row=7 + i, column=1).value = pceen.nom
        sheet.cell(row=7 + i, column=2).value = pceen.prenom
        sheet.cell(row=7 + i, column=3).value = pceen.promo

        sheet.cell(row=7 + i, column=4).value = pceen.email
        if pceen.email and "@" in pceen.email:
            sheet.cell(row=7 + i, column=4).hyperlink = f"mailto:{pceen.email}"
            sheet.cell(row=7 + i, column=4).style = "Hyperlink"

        placesattr = voeu.places_attribuees or 0
        sheet.cell(row=7 + i, column=5).value = placesattr
        npa += placesattr

        apayer = placesattr * spec.unit_price
        sheet.cell(row=7 + i, column=6).value = apayer
        sheet.cell(row=7 + i, column=6).number_format = "#,##0.00 €_-"
        tot += apayer

    tab = openpyxl.worksheet.table.Table(displayName="Données", ref=f"A6:F{7 + i}")

    style = openpyxl.worksheet.table.TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    sheet.cell(row=7 + i + 2, column=1).value = "Total"
    sheet.cell(row=7 + i + 2, column=1).font = openpyxl.styles.Font(bold=True)

    sheet.cell(row=7 + i + 2, column=5).value = npa
    sheet.cell(row=7 + i + 2, column=5).font = openpyxl.styles.Font(bold=True)

    sheet.cell(row=7 + i + 2, column=6).value = tot
    sheet.cell(row=7 + i + 2, column=6).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=7 + i + 2, column=6).number_format = "#,##0.00 €_-"

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 35
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 14

    # Save the Excel workbook to a BytesIO buffer
    excel_data = io.BytesIO()
    workbook.save(excel_data)

    return excel_data


def export_pdf_spectacles(spectacles, season, voeux_attrib_list, voeux_nan_attrib_list):

    # Create an in-memory zip file
    zip_buffer = io.BytesIO()

    # Create a ZipFile object
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        # Add PDF files to the zip archive
        i = 0
        for spectacle in spectacles:  # Replace with your logic to get PDF files
            pdf_data = pdf_spectacle(spectacle, season, voeux_attrib_list[i], voeux_nan_attrib_list[i]).read()
            zipf.writestr(f"{spectacle.nom}.pdf", pdf_data)
            i += 1

    # Move the zip file's pointer to the beginning
    zip_buffer.seek(0)

    return zip_buffer


def export_excel_spectacles(spectacles, season, voeux_attrib_list, voeux_nan_attrib_list):

    # Create an in-memory zip file
    zip_buffer = io.BytesIO()

    # Create a ZipFile object
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
        # Add excel files to the zip archive
        i = 0
        for spectacle in spectacles:  # Replace with your logic to get excel files
            excel_data = excel_spectacle(spectacle, season, voeux_attrib_list[i], voeux_nan_attrib_list[i]).getvalue()
            zipf.writestr(f"{spectacle.nom}.xlsx", excel_data)
            i += 1

    # Move the zip file's pointer to the beginning
    zip_buffer.seek(0)

    return zip_buffer


def exporter_excel_prix(saison, pceens, spectacles, voeux):
    # Code de génération de résumé des payements au Club Q originalement par Loïc Simon - Adapté aux besoin de pem
    # Code assez (très) moche, attention - il y a moyen de faire ça beaucoup mieux, mais bon...
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Récapitulatif"

    sheet["A1"] = f"Club Q – Récapitulatif saison {saison.nom}"
    sheet["A1"].font = openpyxl.styles.Font(bold=True)

    sheet["A2"] = "Début :"
    sheet["B2"] = saison.debut
    sheet["B2"].number_format = "dd/mm/yyyy"
    sheet["D2"] = "Fin :"
    sheet["E2"] = saison.fin
    sheet["E2"].number_format = "dd/mm/yyyy"

    sheet["A3"] = "Spectacles :"
    sheet["B3"] = len(spectacles)
    sheet["D3"] = "Inscrits :"
    sheet["E3"] = pceens.count()

    sheet["A4"] = "Places proposées :"
    sheet["B4"] = sum((spec.nb_tickets or 0) for spec in spectacles)
    sheet["D4"] = "Places vendues :"
    sheet["E4"] = sum((voeu.places_attribuees or 0) for voeu in voeux)

    sheet["A6"] = "NOM"
    sheet["B6"] = "Prénom"
    sheet["C6"] = "Promo/autre"
    sheet["D6"] = "Adresse mail"
    sheet["E6"] = "Places"
    sheet["F6"] = "Somme due"
    sheet["G6"] = "Payé ?"

    npa = 0
    tot = 0
    for i, client in enumerate(sorted(pceens, key=lambda c: c.full_name)):
        sheet.cell(row=7 + i, column=1).value = client.nom.upper()
        sheet.cell(row=7 + i, column=2).value = client.prenom.title()
        sheet.cell(row=7 + i, column=3).value = client.promo or client.autre

        sheet.cell(row=7 + i, column=4).value = client.email
        if client.email and "@" in client.email:
            sheet.cell(row=7 + i, column=4).hyperlink = f"mailto:{client.email}"
            sheet.cell(row=7 + i, column=4).style = "Hyperlink"

        placesattr = pceen_sum_places_attribuees(client, voeux)
        sheet.cell(row=7 + i, column=5).value = placesattr
        npa += placesattr
        apayer = pceen_prix_total(client, voeux)
        sheet.cell(row=7 + i, column=6).value = apayer
        sheet.cell(row=7 + i, column=6).number_format = "#,##0.00 €_-"
        tot += apayer

    tab = openpyxl.worksheet.table.Table(displayName="Données", ref=f"A6:G{7 + i}")

    style = openpyxl.worksheet.table.TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    sheet.cell(row=7 + i + 2, column=1).value = "Total"
    sheet.cell(row=7 + i + 2, column=1).font = openpyxl.styles.Font(bold=True)

    sheet.cell(row=7 + i + 2, column=5).value = npa
    sheet.cell(row=7 + i + 2, column=5).font = openpyxl.styles.Font(bold=True)

    sheet.cell(row=7 + i + 2, column=6).value = tot
    sheet.cell(row=7 + i + 2, column=6).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=7 + i + 2, column=6).number_format = "#,##0.00 €_-"

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 35
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 14
    sheet.column_dimensions["G"].width = 12

    # Save the Excel workbook to a BytesIO buffer
    excel_data = io.BytesIO()
    workbook.save(excel_data)

    return excel_data


def pceen_sum_places_demandees(pceen, voeux) -> int:
    """Gives the number of wanted places for a pceen for the given saison of Club Q"""
    voeux = voeux.filter_by(_pceen_id=pceen.id).all()
    return sum(voeu.places_demandees for voeu in voeux)


def pceen_sum_places_attribuees(pceen, voeux) -> int:
    """Gives the number of given places for a pceen for the given saison of Club Q"""
    voeux = voeux.filter_by(_pceen_id=pceen.id).all()
    return sum(voeu.places_attribuees for voeu in voeux)


def pceen_prix_total(pceen, voeux) -> int:
    """Gives the the total price for a pceen for the given saison of Club Q"""
    voeux = voeux.filter_by(_pceen_id=pceen.id).all()
    return sum(voeu.places_attribuees * voeu.spectacle.unit_price for voeu in voeux)


def sum_object(subject, query) -> int:
    """Return the sum of object in the query corresponding to the subject"""
    if type(subject) == PCeen:
        return query.filter_by(_pceen_id=subject.id).count()
    elif type(subject) == ClubQSpectacle:
        return query.filter_by(_spectacle_id=subject.id).count()
    elif type(subject) == ClubQSalle:
        return query.filter_by(_salle_id=subject.id).count()
    elif type(subject) == ClubQSeason:
        return query.filter_by(_season_id=subject.id).count()


def spectacles_sum_places_attribuees(spectacles) -> int:
    """Gives the number of given places for a saison of Club Q"""
    return sum(spectacle.sum_places_attribuees for spectacle in spectacles)


def spectacles_sum_places_demandees(spectacles) -> int:
    """Gives the number of wanted places for a saison of Club Q"""
    return sum(spectacle.sum_places_demandees for spectacle in spectacles)


def spectacles_sum_places(spectacles) -> int:
    """Gives the total numbers of tickets for the given saison of Club Q"""
    return sum(spectacle.nb_tickets for spectacle in spectacles)


def voeu_form(spectacles, spectacle, pceens, pceen, season_id, redirect, can_edit):

    if spectacle != None:
        # Add spectacle select choice for adding voeu
        setattr(
            forms.AddVoeu,
            "spectacle_add",
            wtforms.SelectField(
                _("Spectacle"),
                choices=[[spect.id, spect.nom] for spect in spectacles],
                default=spectacle.id,
                validators=[DataRequired()],
            ),
        )
    else:
        # Add spectacle select choice for adding voeu
        setattr(
            forms.AddVoeu,
            "spectacle_add",
            wtforms.SelectField(
                _("Spectacle"), choices=[[spect.id, spect.nom] for spect in spectacles], validators=[DataRequired()]
            ),
        )

    # Add pceens select choice for adding voeu
    setattr(
        forms.AddVoeu,
        "pceen_add",
        wtforms.SelectField(
            _("Pcéen"),
            choices=[[pceen_.id, pceen_.full_name] for pceen_ in pceens],
            default=pceen.id,
            validators=[DataRequired()],
        ),
    )

    # Forms
    form_add_voeu = forms.AddVoeu()
    form_edit_voeu = forms.EditVoeu()

    # If a form is validated
    if form_add_voeu.is_submitted() and can_edit:
        submit = form_add_voeu["submit_add"].data
        if submit:
            if form_add_voeu.validate():
                voeu = ClubQVoeu()
                voeu._spectacle_id = form_add_voeu["spectacle_add"].data
                voeu._pceen_id = form_add_voeu["pceen_add"].data
                voeu._season_id = season_id
                voeu.priorite = form_add_voeu["priorite_add"].data
                voeu.places_demandees = form_add_voeu["places_demandees_add"].data
                voeu.places_minimum = form_add_voeu["places_minimum_add"].data or 0
                voeu.places_attribuees = form_add_voeu["places_attribuees_add"].data

                db.session.add(voeu)
                db.session.commit()
                flask.flash(_("Voeu ajouté."))
                return flask.redirect(redirect)

    if form_edit_voeu.is_submitted() and can_edit:
        delete = form_edit_voeu["delete_edit"].data
        if delete:
            voeu = ClubQVoeu.query.filter_by(id=form_edit_voeu["id_edit"].data).one()
            db.session.delete(voeu)

            db.session.commit()
            flask.flash(_("Voeu supprimé."))
            return flask.redirect(redirect)

        if form_edit_voeu.validate():
            voeu = ClubQVoeu.query.filter_by(id=form_edit_voeu["id_edit"].data).one()

            voeu.priorite = form_edit_voeu["priorite_edit"].data
            voeu.places_demandees = form_edit_voeu["places_demandees_edit"].data
            voeu.places_minimum = form_edit_voeu["places_minimum_edit"].data or 0
            voeu.places_attribuees = form_edit_voeu["places_attribuees_edit"].data

            db.session.commit()
            flask.flash(_("Voeu édité."))
            return flask.redirect(redirect)


def spectacle_form(salles, salle, season_id, redirect, can_edit):

    if salle != None:
        setattr(
            forms.EditSpectacle,
            "salle_id",
            wtforms.SelectField(
                _("Salle"),
                choices=[[salle.id, salle.nom] for salle in salles],
                default=salle.id,
                validators=[DataRequired()],
            ),
        )
    else:
        setattr(
            forms.EditSpectacle,
            "salle_id",
            wtforms.SelectField(
                _("Salle"), choices=[[salle.id, salle.nom] for salle in salles], validators=[DataRequired()]
            ),
        )

    form_spectacle = forms.EditSpectacle()

    if form_spectacle.validate_on_submit() and can_edit:
        add = form_spectacle["add"].data

        if add:
            spectacle = ClubQSpectacle()

            db.session.add(spectacle)

            spectacle.nom = form_spectacle["nom"].data
            spectacle._season_id = season_id
            spectacle._salle_id = form_spectacle["salle_id"].data
            spectacle.description = form_spectacle["description"].data
            spectacle.categorie = form_spectacle["categorie"].data
            spectacle.date = datetime.combine(form_spectacle["date"].data, form_spectacle["time"].data)
            spectacle.nb_tickets = form_spectacle["nb_tickets"].data
            spectacle.unit_price = form_spectacle["price"].data

            db.session.commit()

            if form_spectacle["image"].data is not None:
                path = os.path.join(flask.current_app.config["CLUB_Q_BASE_PATH"], str(spectacle.season.id))
                save_path = os.path.join(path, str(spectacle.id) + ".jpg")
                if not os.path.exists(path):
                    os.makedirs(path)
                if os.path.exists(save_path):
                    os.remove(save_path)
                form_spectacle["image"].data.save(save_path)

            flask.flash(_("Spectacle ajouté."))
            return flask.redirect(redirect)

        delete = form_spectacle["delete"].data
        spectacle = ClubQSpectacle.query.filter_by(id=form_spectacle["id"].data).one()

        path = os.path.join(flask.current_app.config["CLUB_Q_BASE_PATH"], str(spectacle.season.id))
        save_path = os.path.join(path, str(spectacle.id) + ".jpg")

        if delete:
            try:
                os.remove(save_path)
            except FileNotFoundError:
                pass

            db.session.delete(spectacle)
            db.session.commit()

            flask.flash(_("Spectacle supprimé."))
            return flask.redirect(redirect)

        else:
            spectacle.nom = form_spectacle["nom"].data
            spectacle._salle_id = form_spectacle["salle_id"].data
            spectacle.description = form_spectacle["description"].data
            spectacle.categorie = form_spectacle["categorie"].data
            spectacle.date = datetime.combine(form_spectacle["date"].data, form_spectacle["time"].data)
            spectacle.nb_tickets = form_spectacle["nb_tickets"].data
            spectacle.unit_price = form_spectacle["price"].data

            if form_spectacle["image"].data is not None:
                if not os.path.exists(path):
                    os.makedirs(path)
                if os.path.exists(save_path):
                    os.remove(save_path)
                form_spectacle["image"].data.save(save_path)

            db.session.commit()
            flask.flash(_("Spectacle édité."))
            return flask.redirect(redirect)
