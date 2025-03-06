import datetime
import openpyxl
import io
import flask

from jours_feries_france import JoursFeries
from flask_babel import _

app = flask.Flask(__name__)


def command_open(panier_bio_day, today: datetime.datetime, day_to_command: datetime.datetime):
    """
    Returns True if according to current date, the command for a date is opened, else returns False.
    Commands are close between Sunday 23h00 and Wednesday 00h01.
    """
    hour_close = 14 #Close at 14h the day before
    daytime_to_command = datetime.datetime.combine(day_to_command, datetime.time(hour=hour_close)) #The day we try to check if opened

    if daytime_to_command - today < datetime.timedelta(days=7):
        days_to_day = today.isoweekday() - panier_bio_day
        if days_to_day <= 0:
            days_to_day = abs(days_to_day)
        else:
            days_to_day = 7 - days_to_day

        if days_to_day < 1:  # If we are one the panier bio day
            return False
        elif days_to_day == 1 and today.hour >= hour_close:  # If we are the day before but after hour_close
            return False
        else:
            return True
    else:
        return(True)


def what_are_next_days(panier_bio_day: int, today: datetime.date, periods):
    """
    Return the date of the next panier bio day after the today's date in the object datetime.date
    """
    next_days = []
    for period in periods:
        date = period.start_date

        if date.year != period.end_date.year:
            jours_feries = JoursFeries.for_year(date.year) + JoursFeries.for_year(period.end_date.year)
        else:
            jours_feries = JoursFeries.for_year(date.year)

        while date <= period.end_date:
            #Checking if upcoming panier bio day
            if date >= today and date.isoweekday() - panier_bio_day == 0:  # ISO format : Monday is 1, Sunday is 7
                #Checking if holyday
                ferie = False
                for i in jours_feries:
                    if jours_feries[i] == date:
                        ferie = True
                if not ferie:
                    next_days.append(date)

            date += datetime.timedelta(days=1)
    return next_days


def get_period_id(date, all_periods):
    date = datetime.datetime.strptime(date,"%A %d %B %Y").date()
    for period in all_periods:
        if date > period.start_date and date < period.end_date:
            return(period.id)
    flask.abort(404)


def validPeriodDates(form, periods):
    #Check if dates are coherent between themselves
    if form["start_date"].data > form["end_date"].data:
        return(False, _("La date de fin ne peut être antérieure à la date de début"))
    
    #Check if periods are superposing between themselves
    for period in periods:
        if period.id == form["id"].data:
            pass

        else:
            if form["start_date"].data < period.end_date and form["end_date"].data > period.start_date:
                return(False, "Les dates choisies se superposent à une période existante")
            
    return(True, "")


def findPeriodId(date, periods):
    """
    Check if a given date fits inside a period
    """
    for period in periods:
        if date <= period.end_date and date >= period.start_date:
            return(period.id)
    else:
        return None

def export_excel(date, orders):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Panier Bio"

    sheet["A1"] = f"Panier Bio – Récapitulatif {date}"
    sheet["A1"].font = openpyxl.styles.Font(bold=True)

    sheet["A2"] = "Date"
    sheet["B2"] = date

    sheet["A4"] = "NOM"
    sheet["B4"] = "Compte Pcéen"
    sheet["C4"] = "Promo/autre"
    sheet["D4"] = "Téléphone"
    sheet["E4"] = "Date"
    sheet["F4"] = "Méthode de payement"
    sheet["G4"] = "Commentaire"
    sheet["H4"] = "Payé?"
    sheet["I4"] = "Payement validé par le trésorier?"
    sheet["J4"] = "Panier récupéré?"

    tot = 0
    for i, order in enumerate(sorted(orders, key=lambda o: o.name)):

        sheet.cell(row=5 + i, column=1).value = order.name
        sheet.cell(row=5 + i, column=2).value = order._pceen_id if order._pceen_id is not None else None
        sheet.cell(row=5 + i, column=3).value = order.service
        sheet.cell(row=5 + i, column=4).value = order.phone_number
        sheet.cell(row=5 + i, column=5).value = order.date
        sheet.cell(row=5 + i, column=6).value = order.payment_method
        sheet.cell(row=5 + i, column=7).value = order.comment

        sheet.cell(row=5 + i, column=8).value = order.payment_made
        sheet.cell(row=5 + i, column=9).value = order.treasurer_validate
        sheet.cell(row=5 + i, column=10).value = order.taken

        for j in range(3):
            sheet.cell(row=5 + i, column=8 + j).style = "Check Cell"

        tot += 1

    tab = openpyxl.worksheet.table.Table(displayName="Données", ref=f"A4:J{5 + i}")

    style = openpyxl.worksheet.table.TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    sheet.cell(row=5 + i + 2, column=1).value = "Total"
    sheet.cell(row=5 + i + 2, column=1).font = openpyxl.styles.Font(bold=True)

    sheet.cell(row=5 + i + 2, column=2).value = tot
    sheet.cell(row=5 + i + 2, column=2).font = openpyxl.styles.Font(bold=True)

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 14
    sheet.column_dimensions["D"].width = 35
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 14

    # Save the Excel workbook to a BytesIO buffer
    excel_data = io.BytesIO()
    workbook.save(excel_data)

    return excel_data


